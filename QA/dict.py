import openpyxl
import random

# 讀取excel sheet
xlsx = openpyxl.load_workbook('dict.xlsx')
sheet = xlsx['dict']
# 第一題
topic_1_q = sheet.cell(column=1, row=1).value
topic_1_a = sheet.cell(column=2, row=1).value
topic_1_b = sheet.cell(column=3, row=1).value
topic_1_c = sheet.cell(column=4, row=1).value
topic_1_d = sheet.cell(column=5, row=1).value
# 第二題
topic_2_q = sheet.cell(column=1, row=2).value
topic_2_a = sheet.cell(column=2, row=2).value
topic_2_b = sheet.cell(column=3, row=2).value
topic_2_c = sheet.cell(column=4, row=2).value
topic_2_d = sheet.cell(column=5, row=2).value
# 第三題
topic_3_q = sheet.cell(column=1, row=3).value
topic_3_a = sheet.cell(column=2, row=3).value
topic_3_b = sheet.cell(column=3, row=3).value
topic_3_c = sheet.cell(column=4, row=3).value
topic_3_d = sheet.cell(column=5, row=3).value
# 第四題
topic_4_q = sheet.cell(column=1, row=4).value
topic_4_a = sheet.cell(column=2, row=4).value
topic_4_b = sheet.cell(column=3, row=4).value
topic_4_c = sheet.cell(column=4, row=4).value
topic_4_d = sheet.cell(column=5, row=4).value
# 第五題
topic_5_q = sheet.cell(column=1, row=5).value
topic_5_a = sheet.cell(column=2, row=5).value
topic_5_b = sheet.cell(column=3, row=5).value
topic_5_c = sheet.cell(column=4, row=5).value
topic_5_d = sheet.cell(column=5, row=5).value
# 第六題
topic_6_q = sheet.cell(column=1, row=6).value
topic_6_a = sheet.cell(column=2, row=6).value
topic_6_b = sheet.cell(column=3, row=6).value
topic_6_c = sheet.cell(column=4, row=6).value
topic_6_d = sheet.cell(column=5, row=6).value
# 第七題
topic_7_q = sheet.cell(column=1, row=7).value
topic_7_a = sheet.cell(column=2, row=7).value
topic_7_b = sheet.cell(column=3, row=7).value
topic_7_c = sheet.cell(column=4, row=7).value
topic_7_d = sheet.cell(column=5, row=7).value
# 第八題
topic_8_q = sheet.cell(column=1, row=8).value
topic_8_a = sheet.cell(column=2, row=8).value
topic_8_b = sheet.cell(column=3, row=8).value
topic_8_c = sheet.cell(column=4, row=8).value
topic_8_d = sheet.cell(column=5, row=8).value
# 第九題
topic_9_q = sheet.cell(column=1, row=9).value
topic_9_a = sheet.cell(column=2, row=9).value
topic_9_b = sheet.cell(column=3, row=9).value
topic_9_c = sheet.cell(column=4, row=9).value
topic_9_d = sheet.cell(column=5, row=9).value
# 第十題
topic_10_q = sheet.cell(column=1, row=10).value
topic_10_a = sheet.cell(column=2, row=10).value
topic_10_b = sheet.cell(column=3, row=10).value
topic_10_c = sheet.cell(column=4, row=10).value
topic_10_d = sheet.cell(column=5, row=10).value

# [[題目一],[題目N]]
topic = []
topic.append([topic_1_q, topic_1_a, topic_1_b, topic_1_c, topic_1_d])
topic.append([topic_2_q, topic_2_a, topic_2_b, topic_2_c, topic_2_d])
topic.append([topic_3_q, topic_3_a, topic_3_b, topic_3_c, topic_3_d])
topic.append([topic_4_q, topic_4_a, topic_4_b, topic_4_c, topic_4_d])
topic.append([topic_5_q, topic_5_a, topic_5_b, topic_5_c, topic_5_d])
topic.append([topic_6_q, topic_6_a, topic_6_b, topic_6_c, topic_6_d])
topic.append([topic_7_q, topic_7_a, topic_7_b, topic_7_c, topic_7_d])
topic.append([topic_8_q, topic_8_a, topic_8_b, topic_8_c, topic_8_d])
topic.append([topic_9_q, topic_9_a, topic_9_b, topic_9_c, topic_9_d])
topic.append([topic_10_q, topic_10_a, topic_10_b, topic_10_c, topic_10_d])

# 打亂題目list
topic_random = random.sample(topic, len(topic))
